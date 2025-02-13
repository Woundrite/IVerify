from openpyxl import load_workbook
import datetime as dt
import os
import xlsxwriter
import json
import asyncio
from pyppeteer import launch
import platform
from xlsx2html import xlsx2html
import sys
import tkinter as tk
from Popup import Popup

isDebugging = False

std_out = sys.stdout
file_out = open("log.txt", "w")
sys.stdout = file_out
def HandleError(e, root):
	Popup(title="Error", message=f"An error occurred: {str(e)}", master=root)
	print(e)
	# sys.stdout = std_out
	# file_out.close()
	# sys.exit(1)
	
async def generate_pdf(url, pdf_path, tk_root):
	try:
		bpth = os.getcwd()+'/chrome-win/chrome.exe'
		print(bpth)
		if platform.system() != "Windows":
			bpth = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"
		browser = await launch(headless=True, executablePath=bpth, options={'args': ['--no-sandbox']})
		page = await browser.newPage()
		
		await page.goto(url)
		print(f"Generated PDF: {pdf_path}")
		await page.pdf({'path': pdf_path, 'format': 'A4', 'printBackground': True})
	except Exception as e:
		print(f"Error generating PDF: {e}")
	finally: 	
		await browser.close()
	# converter.convert(f'{url}', f'{pdf_path}')

def to_pdf(file_path, in_root, tk_root):
	print(f"Generating PDF for of {file_path}...")
	file_path = file_path.replace('\\', '/')
	os.path.relpath(file_path, os.getcwd()) 
	print(file_path)
	xlsx2html(file_path, 'file.html')
	html = ""
	with open('file.html', 'r') as file:
		html = file.read()
	html = html.replace("""</body>""", """<div style="width: 90vw; display:grid; justify-content: center; padding-top: 5px;">
		<a href="https://unifiedportal-epfo.epfindia.gov.in">https://unifiedportal-epfo.epfindia.gov.in</a>
	</div>
</body>""")
	
	html = html.replace("""<body>""", f"""<body>
	<div style="width: 95vw; display:grid; padding-top: 5px; grid-template-columns: 1fr auto auto;">
		<div></div>
		<p style="width: 10vw">Date/Time</p>
		<p>{dt.datetime.now().strftime("%d-%b-%y %I:%M %p")}</p>
	</div>
""")
	with open('file.html', 'w') as file:
		file.write(html)
	asyncio.run(generate_pdf(f'file:///{os.getcwd()+'/file.html'}', file_path.replace('.xlsx', '.pdf'), tk_root))

def handle_excel_write(file_path, personal_data, tk_root):
	# Create a workbook and add a worksheet.
	workbook = xlsxwriter.Workbook(file_path)
	worksheet = workbook.add_worksheet()
	# Add a bold format to use to highlight cells.
	bold = workbook.add_format({'bold': True})
	# add a blue background to the headers.
	header_format = workbook.add_format({'bold': True, 'bg_color': '#95b3d7', 'align': 'center', 'valign': 'vcenter', 'border': 1})
	# add border to the cells.
	cell_format = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter'})
	worksheet.write('A1', 'UAN NUMBER', header_format)
	worksheet.write('B1', 'MEMBER ID', header_format)
	worksheet.write('C1', 'ESTABLISHMENT DETAILS', header_format)
	worksheet.write('D1', 'NAME', header_format)
	worksheet.write('E1', 'FATHER OR HUSBAND \n NAME', header_format)
	worksheet.write('F1', 'DATE OF JOIN', header_format)
	worksheet.write('G1', 'DATE OF EXIT PF', header_format)
	last_row = 1
	# Start from the first cell below the headers.
	for row, index in enumerate(personal_data): # has the following format: [str(uan), row['MemID'], est, Name, FName, dt.datetime.strptime(row["EPFDOJ"], '%d-%b-%Y'), ""]
		for col, item in enumerate(index):
			if col >= 5:
				if type(item) == str:
					worksheet.write(row + 1, col, item, cell_format)
				else:
					worksheet.write(row + 1, col, item.strftime('%d-%b-%y'), cell_format)
			else:
				worksheet.write(row + 1, col, item, cell_format)
		last_row += 1
	worksheet.autofit()

	workbook.close()
	


def handle_input(file_path, uan, root, uan_name_map, company_names_map, tk_root):
	if uan == None:
		return True, None
	output = []
	workbook = load_workbook(filename=file_path)
	sheet = workbook.active

	personal_data = {}

	# Using the values_only because you want to return the cells' values
	for row in sheet.iter_rows(min_row=4,
							min_col=0,
							max_col=14,
							values_only=True):
		personal_data.update({row[0]: {'MemID': row[0], 
									   'EstID': row[1],
									   'EPFDOJ': row[2], 
									   'EPSDOJ': row[3],
									   'EPFDOE': row[4],
									   'EPSDOE': row[5],
									   'ROL': row[6],
									   'PE': row[7], 
									   'BT': row[8],
									   'CS': row[9],
									   'BalSt': row[10],
									   'NEMem': row[11],
									   'UANLink': row[12],}})
	if len(personal_data) == 0:
		if not isDebugging:
			HandleError(f"No data found in file {file_path}", tk_root)
		return
	
	print(personal_data)

	EKeys = [] # keys of EstID not found
	for row in personal_data.values():
		name = uan_name_map.get(uan, uan)
		if type(name) == dict:
			FName = name.get("FName", "").upper()
			Name = name.get("Name", "").upper()
		else:
			if not isDebugging:
				HandleError(f"No Name and Father's name provided for UAN {uan}", tk_root)
				continue
			FName = ""
			Name = name
			
		try:
			est = company_names_map[row["EstID"]]
		except KeyError:
			print(f"Could not find company name for EstID {row['EstID']}")
			EKeys.append(row["EstID"])
			est = ""
		print(output)

		print(row)
		
		try:
			if type(row["EPFDOE"]) in [str, type(None)]:
				if row["EPFDOE"] == None or row["EPFDOE"].strip() == "":
					print(f"No date of exit provided for UAN {uan}")
					row["EPFDOE"] = ""
				else:
					row["EPFDOE"] = dt.datetime.strptime(row["EPFDOE"], '%Y-%m-%d')
			else:
				row["EPFDOE"] = ""
				# if not isDebugging:
				# 	HandleError(f"Invalid DOE format for UAN {uan}", tk_root)
				# 	print(f"Invalid DOE format for UAN {uan}", tk_root)

			if  type(row["EPFDOJ"]) in [str, type(None)]:

				if row["EPFDOJ"] == None or row["EPFDOJ"].strip() == "":
					row["EPFDOJ"] = ""
					print(f"No date of join provided for UAN {uan}")
				else:
					row["EPFDOJ"] = dt.datetime.strptime(row["EPFDOJ"], '%Y-%m-%d')
			else:
				row["EPFDOJ"] = ""
				# if not isDebugging:
				# 	HandleError(f"Invalid DOJ format for UAN {uan}", tk_root)
				# 	print(f"Invalid DOJ format for UAN {uan}", tk_root)
		except Exception as e:
			print(f"Error parsing date for UAN {uan}: {e}")
			print(file_path, uan, root, row)
			sys.exit(1)
		
		output.append([str(uan), row['MemID'], est, Name, FName, row["EPFDOJ"], row["EPFDOE"]])
	if len(EKeys) > 0:
		return True, EKeys
	print(output)
	output = sorted(output, key=lambda x: x[5] if x[5] != "" else dt.datetime.now())[::-1]
	handle_excel_write(os.path.join(root, f"{uan}.xlsx"), output, tk_root)
	workbook.close()
	return False, output

def compute(company_name_map_file, uan_id_compiled_file, in_root, out_root, tk_root):
	print(f"Company Name Map: {company_name_map_file}")
	print(f"UAN ID Compiled File: {uan_id_compiled_file}")
	print(f"Input Root: {in_root}")
	print(f"Output Root: {out_root}")
	if not os.path.exists(out_root):
		os.makedirs(out_root)	
	company_name_workbook = load_workbook(filename=company_name_map_file)
	company_name_sheet = company_name_workbook.active
	company_names_map = {}

	# Using the values_only because you want to return the cells' values
	for row in company_name_sheet.iter_rows(min_row=2,
							min_col=0,
							max_col=2,
							values_only=True):
		company_names_map.update({row[0]:  row[1]})

	print(json.dumps(company_names_map, indent=2))
	company_name_workbook.close()


	uan_name_workbook = load_workbook(filename=uan_id_compiled_file)
	uan_name_map = {}
	uan_id_map = []
	uan_name_sheet = uan_name_workbook.active
	for row in uan_name_sheet.iter_rows(min_row=0,
							min_col=0,
							max_col=3,
							values_only=True):
		if row[0] != None:
			if row[2] in ["", " ", None]:
				HandleError(f"Father's Name not found for UAN ID {row[0]}", tk_root)
				continue
			if row[1] in ["", " ", None]:
				HandleError(f"Name not found for UAN ID {row[0]}", tk_root)
				continue
		uan_name_map.update({str(row[0]): {'Name': row[1], 'FName': row[2]}})
		uan_id_map.append(row[0])
	uan_name_workbook.close()

	# uan_id_workbook = load_workbook(filename=uan_id_compiled_file)
	# uan_id_sheet = uan_id_workbook.active
	# for row in uan_id_sheet.iter_rows(min_row=2,
	# 						min_col=0,
	# 						max_col=2,
	# 						values_only=True):
		
	# 	uan_id_map.append(row[0])
	# 	if uan_name_map.get(str(row[0]), None) is not None:
	# 		# if the Fname column is not empty, use it as the Father's name
	# 		if uan_name_map[str(row[0])]["FName"] in ["", " ", None]:
	# 			print("NAME MAP!!!", row)
	# 			uan_name_map[str(row[0])] = {'Name': row[1], 'FName': uan_name_map[str(row[0])]["Name"]}

	print("\n\n UAN Name MAP\n\n", json.dumps(uan_name_map, indent=2))
	# print(json.dumps(uan_id_map, indent=2))
	# uan_id_workbook.close()



	data = []
	res_comp = []
	for root, dirs, files in os.walk(in_root):
		path = root.split(os.sep)
		for file in files:
			print("Found File: ", file)
			if "uanacord_uan_member_data" in file:
				if file[-17:-5] in uan_name_map.keys():
					print(f"Processing {file}...")
				else:
					print(f"UAN ID not found for {file}...")
				
				error, res = handle_input(os.path.join(root, file), file[-17:-5], out_root, uan_name_map, company_names_map, tk_root)
				if error:
					if res != None:
						res_comp.extend(res)
				else:
					print(res)
					data.append(res)

	if len(res_comp) > 0:
		# for filename in os.listdir(out_root):
		# 	file_path = os.path.join(out_root, filename)
		# 	try:
		# 		if os.path.isfile(file_path):
		# 			os.remove(file_path)
		# 	except Exception as e:
		# 		print('Failed to delete %s. Reason: %s' % (file_path, e))
		if not isDebugging:
			comps = "\n".join([", ".join(res_comp[x:x+5]) for x in range(0, len(data), 5)])
			HandleError(f"No company name(s) found for EstID: {comps}", tk_root)
	
	# convert all xlsx files to pdf
	for root, dirs, files in os.walk(out_root):
		for file in files:
			if file.endswith(".xlsx"):
				to_pdf(os.path.join(root, file), in_root, tk_root)
	
	combine_xlsx(out_root, data, tk_root)

	sys.stdout = std_out
	file_out.close()

def combine_xlsx(out_root, data, tk_root):
	print("Combining xlsx files...")
	print("\n\n")
	# reads all the xlsx files in the directory and combines all the data into one xlsx file and saves it, uses xlsx writer
	# seperate tables for each uan id
	# Create a workbook and add a worksheet.
	workbook = xlsxwriter.Workbook(os.path.join(out_root, "Excel Reports.xlsx"))
	worksheet = workbook.add_worksheet()
	# header's format.
	header_format = workbook.add_format({'bold': True, 'bg_color': '#95b3d7','align': 'center', 'valign': 'vcenter', 'border': 1})
	# add border to the cells.
	cell_format = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter'})
	row = 0
	# print(json.dumps(data, indent=4))
	
	for group in data:
		worksheet.write(row, 0, 'UAN NUMBER', header_format)
		worksheet.write(row, 1, 'MEMBER ID', header_format)
		worksheet.write(row, 2, 'ESTABLISHMENT DETAILS', header_format)
		worksheet.write(row, 3, 'NAME', header_format)
		worksheet.write(row, 4, 'FATHER OR HUSBAND \n NAME', header_format)
		worksheet.write(row, 5, 'DATE OF JOIN', header_format)
		worksheet.write(row, 6, 'DATE OF EXIT PF', header_format)
		row+=1

		for r in group:
			worksheet.write(row, 0, r[0], cell_format)
			worksheet.write(row, 1, r[1], cell_format)
			worksheet.write(row, 2, r[2], cell_format)
			worksheet.write(row, 3, r[3], cell_format)
			worksheet.write(row, 4, r[4], cell_format)

			if type(r[5]) == str:
				worksheet.write(row, 5, r[5], cell_format)
			else:
				worksheet.write(row, 5, r[5].strftime('%d-%b-%y'), cell_format)
			
			if type(r[6]) == str:
				worksheet.write(row, 6, r[6], cell_format)
			else:
				worksheet.write(row, 6, r[6].strftime('%d-%b-%y'), cell_format)

			print("")
			row += 1
		print("\n\n")
	worksheet.autofit()

	workbook.close()
	os.remove("file.html")
	print("Excel Reports.xlsx has been created.")

if __name__ == "__main__":
	company_name_map_file = "C:/Users/ndben/Desktop/Nikhil/Freelancing/IVerify/src/Tests/re77uan/ESTABLIBSHMENT LIST.xlsx"
	# uan_id_compiled_file = "C:/Users/ndben/Desktop/Nikhil/Freelancing/IVerify/src/Tests/T2/Name.xlsx"
	uan_name_compiled_file = "C:/Users/ndben/Desktop/Nikhil/Freelancing/IVerify/src/Tests/re77uan/data.xlsx"
	in_root = "C:/Users/ndben/Desktop/Nikhil/Freelancing/IVerify/src/Tests/re77uan/exl"
	out_root = "C:/Users/ndben/Desktop/Nikhil/Freelancing/IVerify/src/Tests/re77uan/out"
	# company_name_map_file = input("Enter the path to the Company name map file: ")
	# uan_id_compiled_file = input("Enter the path to the UAN ID compiled file: ")
	# uan_name_compiled_file = input("Enter the path to the UAN Name compiled file: ")
	# in_root = input("Enter the base directory path: ")
	# out_root = input("Enter the out directory path: ")
	# Main Application
	root = tk.Tk()
	root.title("Excel and PDF Processor")
	root.geometry("600x600")
	root.resizable(False, False)
	compute(company_name_map_file, uan_name_compiled_file, in_root, out_root, root)
	print("Completed.")